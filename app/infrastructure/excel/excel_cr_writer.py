# app/infrastructure/excel/excel_cr_writer.py
import io
import logging
import openpyxl
from app.application.use_cases.excel_cr.aggregate_and_match import AggregatedRow

logger = logging.getLogger(__name__)


class ExcelCrWriter:
    @staticmethod
    def write(template_bytes: bytes, rows: list[AggregatedRow]) -> bytes:
        """Fill template Excel with aggregated rows. Returns modified Excel bytes."""
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        ws = wb.active

        # Build lookup: chi_tieu_label → row_index
        chi_tieu_col, month_row = _find_header_structure(ws)
        row_index = _build_row_index(ws, chi_tieu_col)
        col_index = _build_col_index(ws, month_row)

        for row in rows:
            if not row.chi_tieu:
                continue
            r = row_index.get(row.chi_tieu.strip())
            c = col_index.get(row.thang)
            if r is None:
                logger.warning("Chi tieu '%s' not found in template — skipping", row.chi_tieu)
                continue
            if c is None:
                logger.warning("Thang %d not found in template — skipping", row.thang)
                continue
            current = ws.cell(row=r, column=c).value or 0
            ws.cell(row=r, column=c).value = current + row.so_tien

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


def _find_header_structure(ws) -> tuple[int, int]:
    """Return (chi_tieu_column_index, month_header_row_index)."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "chỉ tiêu" in str(cell.value).lower():
                return cell.column, cell.row
    return 1, 1


def _build_row_index(ws, chi_tieu_col: int) -> dict[str, int]:
    index = {}
    for row in ws.iter_rows():
        cell = row[chi_tieu_col - 1]
        if cell.value and isinstance(cell.value, str):
            index[cell.value.strip()] = cell.row
    return index


def _build_col_index(ws, header_row: int) -> dict[int, int]:
    """Map month number → column index by scanning header row for 'Tháng N' patterns."""
    import re
    index = {}
    for cell in ws[header_row]:
        if cell.value:
            m = re.search(r"(\d+)", str(cell.value))
            if m:
                month = int(m.group(1))
                index[month] = cell.column
    return index
