import asyncio
import concurrent.futures
from io import BytesIO
from decimal import Decimal
import openpyxl
from app.domain.entities.invoice_item import InvoiceItem
from app.domain.ports.excel_port import IExcelPort

_process_pool = concurrent.futures.ProcessPoolExecutor()

def _append_rows_process(template_bytes: bytes, items: list[InvoiceItem], existing_data: bytes) -> bytes:
    src = existing_data if existing_data else template_bytes
    wb = openpyxl.load_workbook(BytesIO(src))
    ws = wb[SHEET_NAME]

    # Nếu là tháng mới (file chưa có), xóa trắng dữ liệu mẫu trong template
    if not existing_data and ws.max_row >= DATA_START_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)

    # Find next available row after the last data row
    last_row = DATA_START_ROW - 1
    for row in range(ws.max_row, DATA_START_ROW - 1, -1):
        if ws.cell(row=row, column=5).value is not None:  # invoice_number column
            last_row = row
            break

    # Determine STT offset
    stt_offset = 0
    for row in range(DATA_START_ROW, last_row + 1):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, int):
            stt_offset = v

    for idx, item in enumerate(items, start=1):
        r = last_row + idx
        ws.cell(row=r, column=1).value = stt_offset + idx
        ws.cell(row=r, column=4).value = item.invoice_symbol
        ws.cell(row=r, column=5).value = item.invoice_number
        ws.cell(row=r, column=6).value = item.invoice_date
        ws.cell(row=r, column=7).value = item.seller_name
        ws.cell(row=r, column=8).value = item.seller_address
        ws.cell(row=r, column=9).value = item.seller_tax_code
        ws.cell(row=r, column=10).value = item.description
        ws.cell(row=r, column=11).value = float(item.price_before_tax)
        ws.cell(row=r, column=12).value = int(item.tax_rate * 100)
        ws.cell(row=r, column=13).value = float(item.tax_rate)
        ws.cell(row=r, column=14).value = float(item.price_after_tax)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# First data row in the template (after headers and category label)
DATA_START_ROW = 13
SHEET_NAME = "Bang ke thue"

def generate_monthly_filename(month: int, year: int) -> str:
    """Generate filename: Tong_hop_hoa_don_T{month}_{year}.xlsx
    Example: Tong_hop_hoa_don_T4_2026.xlsx
    """
    return f"Tong_hop_hoa_don_T{month}_{year}.xlsx"

class OpenpyxlWriter(IExcelPort):
    def __init__(self, template_path: str):
        self._clean_template_bytes = self._build_clean_template(template_path)

    @staticmethod
    def _build_clean_template(template_path: str) -> bytes:
        """Load template once, strip sample data rows, cache as bytes."""
        wb = openpyxl.load_workbook(template_path)
        ws = wb[SHEET_NAME]
        if ws.max_row >= DATA_START_ROW:
            ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def append_rows(
        self, items: list[InvoiceItem], year: int, month: int, existing_data: bytes
    ) -> tuple[str, bytes]:
        filename = generate_monthly_filename(month, year)
        loop = asyncio.get_running_loop()
        file_bytes = await loop.run_in_executor(
            _process_pool, _append_rows_process, self._clean_template_bytes, items, existing_data
        )
        return filename, file_bytes