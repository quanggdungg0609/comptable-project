import asyncio
import concurrent.futures
from io import BytesIO
from decimal import Decimal
import openpyxl
from app.domain.entities.invoice_line_item import InvoiceLineItem
from app.domain.ports.excel_detail_port import IExcelDetailPort

_process_pool = concurrent.futures.ProcessPoolExecutor()

def _append_rows_detail_process(template_bytes: bytes, items: list[InvoiceLineItem], existing_data: bytes, month: int, year: int) -> bytes:
    src = existing_data if existing_data else template_bytes
    wb = openpyxl.load_workbook(BytesIO(src))
    ws = wb.active

    ws["A1"] = f"Kỳ tính thuế: Tháng {month:02d} năm {year}"

    # Tìm hàng cuối có dữ liệu (cột I = ten_hang_hoa)
    last_row = DATA_START_ROW - 1
    for row in range(ws.max_row, DATA_START_ROW - 1, -1):
        if ws.cell(row=row, column=9).value is not None:
            last_row = row
            break

    # STT offset
    stt_offset = 0
    for row in range(DATA_START_ROW, last_row + 1):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, int):
            stt_offset = v

    for idx, li in enumerate(items, start=1):
        r = last_row + idx
        ws.cell(row=r, column=1).value = stt_offset + idx   # STT
        ws.cell(row=r, column=2).value = ""                  # Ký hiệu mẫu HĐ (trống)
        ws.cell(row=r, column=3).value = li.invoice_symbol   # Ký hiệu HĐ
        ws.cell(row=r, column=4).value = li.invoice_number   # Số HĐ
        ws.cell(row=r, column=5).value = li.invoice_date     # Ngày phát hành
        ws.cell(row=r, column=6).value = li.seller_name      # Tên nhà cung cấp
        ws.cell(row=r, column=7).value = li.seller_address   # Địa chỉ người bán
        ws.cell(row=r, column=8).value = li.seller_tax_code  # MST
        ws.cell(row=r, column=9).value = li.ten_hang_hoa     # Mặt hàng
        ws.cell(row=r, column=10).value = li.don_vi_tinh     # Đơn vị tính
        ws.cell(row=r, column=11).value = float(li.so_luong) # Số lượng
        ws.cell(row=r, column=12).value = float(li.don_gia)  # Đơn giá
        ws.cell(row=r, column=13).value = float(li.thanh_tien)  # Thành tiền
        ws.cell(row=r, column=14).value = int(li.tax_rate * 100)  # Thuế suất %
        ws.cell(row=r, column=15).value = float(li.tax_amount)    # Thuế GTGT

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

DATA_START_ROW = 6  # Hàng data đầu tiên trong template (sau headers)

def generate_detail_filename(month: int, year: int) -> str:
    return f"Chi_tiet_hoa_don_T{month}_{year}.xlsx"

class OpenpyxlDetailWriter(IExcelDetailPort):
    def __init__(self, template_path: str):
        self._clean_template_bytes = self._build_clean_template(template_path)

    @staticmethod
    def _build_clean_template(template_path: str) -> bytes:
        """Load template once, strip sample data rows, cache as bytes."""
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        if ws.max_row >= DATA_START_ROW:
            ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def append_rows(
        self,
        items: list[InvoiceLineItem],
        year: int,
        month: int,
        existing_data: bytes,
    ) -> tuple[str, bytes]:
        filename = generate_detail_filename(month, year)
        loop = asyncio.get_running_loop()
        file_bytes = await loop.run_in_executor(
            _process_pool, _append_rows_detail_process, self._clean_template_bytes, items, existing_data, month, year
        )
        return filename, file_bytes
