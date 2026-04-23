import asyncio
import logging
from io import BytesIO
import openpyxl
from app.domain.ports.storage_port import IStoragePort
from app.domain.ports.job_repository import IJobRepository

logger = logging.getLogger(__name__)

AGGREGATE_SHEET = "Bang ke thue"
AGGREGATE_DATA_START = 13
AGGREGATE_COLS = [1, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
AGGREGATE_HEADERS = [
    "STT", "Ký hiệu HĐ", "Số HĐ", "Ngày HĐ",
    "Tên người bán", "MST", "Diễn giải",
    "Tiền trước thuế", "Thuế suất %", "Hệ số thuế", "Tiền sau thuế",
]

DETAIL_DATA_START = 6
DETAIL_COLS = list(range(1, 15))
DETAIL_HEADERS = [
    "STT", "KH Mẫu HĐ", "Ký hiệu HĐ", "Số HĐ", "Ngày phát hành",
    "Tên NCC", "MST", "Tên hàng hóa", "ĐVT",
    "Số lượng", "Đơn giá", "Thành tiền", "Thuế suất %", "Thuế GTGT",
]

# Storage download timeout (seconds) — avoids hanging when RustFS is slow
_STORAGE_TIMEOUT = 3.0


class GetExportsUseCase:
    def __init__(
        self,
        storage: IStoragePort,
        bucket_exports: str,
        template_aggregate: str,
        template_detail: str,
        repo: IJobRepository,
    ):
        self._storage = storage
        self._bucket_exports = bucket_exports
        self._template_aggregate = template_aggregate
        self._template_detail = template_detail
        self._repo = repo

    async def get_preview(self, year: int, month: int) -> dict:
        logger.info(f"[GetExports] Getting preview for {year}/{month:02d}")
        logger.debug(f"[GetExports] Fetching items and line items from repository")

        items, line_items = await asyncio.gather(
            self._repo.get_items_by_month(year, month),
            self._repo.get_line_items_by_month(year, month),
        )

        logger.info(f"[GetExports] Found {len(items)} items and {len(line_items)} line items for {year}/{month:02d}")

        agg_rows = [
            [
                idx,
                it.invoice_symbol, it.invoice_number, it.invoice_date,
                it.seller_name, it.seller_tax_code, it.description,
                it.price_before_tax, int(it.tax_rate * 100), it.tax_rate, it.price_after_tax,
            ]
            for idx, it in enumerate(items, start=1)
        ]

        detail_rows = [
            [
                idx, "",
                li.invoice_symbol, li.invoice_number, li.invoice_date,
                li.seller_name, li.seller_tax_code, li.ten_hang_hoa, li.don_vi_tinh,
                li.so_luong, li.don_gia, li.thanh_tien,
                int(li.tax_rate * 100), li.tax_amount,
            ]
            for idx, li in enumerate(line_items, start=1)
        ]

        logger.debug(f"[GetExports] Formatted {len(agg_rows)} aggregate rows and {len(detail_rows)} detail rows")

        return {
            "year": year,
            "month": month,
            "aggregate": {"headers": AGGREGATE_HEADERS, "rows": agg_rows},
            "detail": {"headers": DETAIL_HEADERS, "rows": detail_rows},
            "agg_filename": f"Bang_ke_thue_{year}_{month:02d}.xlsx",
            "detail_filename": f"Chi_tiet_hoa_don_T{month}_{year}.xlsx",
        }

    async def get_download(self, year: int, month: int, file_type: str) -> tuple[bytes, str]:
        logger.info(f"[GetExports] Getting download for {file_type} {year}/{month:02d}")

        data = await self._fetch(year, month, file_type)
        filename = (
            f"Bang_ke_thue_{year}_{month:02d}.xlsx"
            if file_type == "aggregate"
            else f"Chi_tiet_hoa_don_T{month}_{year}.xlsx"
        )

        logger.info(f"[GetExports] Returning {filename} ({len(data)} bytes)")
        return data, filename

    async def _fetch(self, year: int, month: int, file_type: str) -> bytes:
        key = (
            f"{year}/{month:02d}/Bang_ke_thue_{year}_{month:02d}.xlsx"
            if file_type == "aggregate"
            else f"{year}/{month:02d}/Chi_tiet_hoa_don_T{month}_{year}.xlsx"
        )

        logger.debug(f"[GetExports] Fetching {file_type} from storage: {key}")

        try:
            logger.debug(f"[GetExports] Attempting to download from storage (timeout: {_STORAGE_TIMEOUT}s)")
            data = await asyncio.wait_for(
                self._storage.download_file(self._bucket_exports, key),
                timeout=_STORAGE_TIMEOUT,
            )
            logger.info(f"[GetExports] Successfully fetched from storage ({len(data)} bytes)")
            return data
        except asyncio.TimeoutError:
            logger.warning(f"[GetExports] Storage download timed out, trying template")
        except Exception as e:
            logger.debug(f"[GetExports] Storage download failed: {e}, trying template")

        # File not in storage — try loading from template, then fall back to minimal xlsx
        try:
            logger.debug(f"[GetExports] Loading from template")
            data = await asyncio.to_thread(self._from_template, file_type)
            logger.info(f"[GetExports] Generated from template ({len(data)} bytes)")
            return data
        except Exception as e:
            logger.warning(f"[GetExports] Template loading failed: {e}, generating minimal xlsx")
            data = self._minimal_empty(file_type)
            logger.info(f"[GetExports] Generated minimal xlsx ({len(data)} bytes)")
            return data

    def _from_template(self, file_type: str) -> bytes:
        template = self._template_aggregate if file_type == "aggregate" else self._template_detail
        wb = openpyxl.load_workbook(template)
        ws = wb[AGGREGATE_SHEET] if file_type == "aggregate" else wb.active
        data_start = AGGREGATE_DATA_START if file_type == "aggregate" else DETAIL_DATA_START
        if ws.max_row >= data_start:
            ws.delete_rows(data_start, ws.max_row - data_start + 1)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _minimal_empty(self, file_type: str) -> bytes:
        """Generate bare-minimum xlsx with just the sheet name — no template needed."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if file_type == "aggregate":
            ws.title = AGGREGATE_SHEET
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _parse_rows(self, data: bytes, file_type: str) -> list[list]:
        try:
            wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        except Exception:
            return []

        try:
            if file_type == "aggregate":
                ws = wb[AGGREGATE_SHEET]
                data_start = AGGREGATE_DATA_START
                cols = AGGREGATE_COLS
            else:
                ws = wb.active
                data_start = DETAIL_DATA_START
                cols = DETAIL_COLS
        except KeyError:
            return []

        rows = []
        for row_idx in range(data_start, ws.max_row + 1):
            row_data = [ws.cell(row=row_idx, column=c).value for c in cols]
            if any(v is not None for v in row_data):
                rows.append(row_data)
        return rows
