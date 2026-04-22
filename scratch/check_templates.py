import openpyxl

def check_template(path, start_row):
    print(f"Checking {path} (Data starts at {start_row})")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    print(f"Max row: {ws.max_row}")
    sample_data = []
    for r in range(start_row, min(start_row + 5, ws.max_row + 1)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        sample_data.append(row_vals)
    
    has_data = any(any(v is not None for v in row) for row in sample_data)
    print(f"Has data in top 5 data rows: {has_data}")
    if has_data:
        for i, rd in enumerate(sample_data):
            print(f" Row {start_row + i}: {rd}")

check_template("Mau_xuat_du_lieu.xlsx", 13)
check_template("Mau_xuat_du_lieu_chi_tiet.xlsx", 6)
