import io
import pytest
import openpyxl
from app.infrastructure.excel.excel_cr_writer import ExcelCrWriter
from app.application.use_cases.excel_cr.aggregate_and_match import AggregatedRow

def _make_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Chỉ tiêu"
    ws["B1"] = "Tháng 1"
    ws["C1"] = "Tháng 2"
    ws["A2"] = "Chi khác"
    ws["A3"] = "Chi phí lao động"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def test_writer_fills_chi_tieu_cell():
    template_bytes = _make_template()
    rows = [
        AggregatedRow(thang=1, dien_giai="Thưởng sáng kiến", khoan_muc="cpk",
                      so_tien=70_900_000.0, chi_tieu="Chi khác", match_tier="llm_confirmed"),
    ]
    result = ExcelCrWriter.write(template_bytes, rows)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    assert ws["B2"].value == pytest.approx(70_900_000.0)

def test_writer_unknown_chi_tieu_is_skipped():
    template_bytes = _make_template()
    rows = [
        AggregatedRow(thang=1, dien_giai="Unknown", khoan_muc="cpk",
                      so_tien=1_000.0, chi_tieu="Không tồn tại", match_tier="direct"),
    ]
    result = ExcelCrWriter.write(template_bytes, rows)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    # B2 and B3 should be None (untouched)
    assert ws["B2"].value is None
    assert ws["B3"].value is None
