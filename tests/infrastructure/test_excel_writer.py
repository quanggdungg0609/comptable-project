import pytest
from decimal import Decimal
from datetime import date
from io import BytesIO
import openpyxl
from app.domain.entities.invoice_item import InvoiceItem
from app.infrastructure.excel.openpyxl_writer import OpenpyxlWriter

TEMPLATE_PATH = "Mau_xuat_du_lieu.xlsx"

def make_item(**kwargs):
    defaults = dict(
        invoice_symbol="1C26TAA", invoice_number="49",
        invoice_date=date(2026, 3, 12), seller_name="Cty XYZ",
        seller_tax_code="0901212659", description="Vật tư",
        price_before_tax=Decimal("29030000"), tax_rate=Decimal("0.10"),
        price_after_tax=Decimal("2903000"),
    )
    defaults.update(kwargs)
    return InvoiceItem(**defaults)

@pytest.mark.asyncio
async def test_append_rows_returns_filename_and_bytes():
    writer = OpenpyxlWriter(template_path=TEMPLATE_PATH)
    filename, file_bytes = await writer.append_rows([make_item()], year=2026, month=3, existing_data=b"")
    assert filename == "Tong_hop_hoa_don_T3_2026.xlsx"
    assert isinstance(file_bytes, bytes)
    assert len(file_bytes) > 0

@pytest.mark.asyncio
async def test_appended_row_contains_invoice_number():
    writer = OpenpyxlWriter(template_path=TEMPLATE_PATH)
    filename, file_bytes = await writer.append_rows([make_item(invoice_number="TEST-49")], year=2026, month=3, existing_data=b"")
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb["Bang ke thue"]
    values = [ws.cell(row=r, column=5).value for r in range(1, ws.max_row + 1)]
    assert "TEST-49" in values