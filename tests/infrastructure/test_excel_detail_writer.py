import pytest
from decimal import Decimal
from datetime import date
from io import BytesIO
import openpyxl
from app.domain.entities.invoice_line_item import InvoiceLineItem
from app.infrastructure.excel.openpyxl_detail_writer import OpenpyxlDetailWriter

TEMPLATE_PATH = "Mau_xuat_du_lieu_chi_tiet.xlsx"

def make_line_item(**kwargs):
    defaults = dict(
        invoice_symbol="1C26TAA", invoice_number="49",
        invoice_date=date(2026, 3, 12),
        seller_name="Cty TNHH ĐT và TM Linh Chi Nguyễn",
        seller_tax_code="0901212659",
        ten_hang_hoa="Thép tấm 10mm", don_vi_tinh="Kg",
        so_luong=Decimal("298"), don_gia=Decimal("28000"),
        thanh_tien=Decimal("8344000"), tax_rate=Decimal("0.10"),
        tax_amount=Decimal("834400"),
    )
    defaults.update(kwargs)
    return InvoiceLineItem(**defaults)

@pytest.mark.asyncio
async def test_append_returns_filename_and_bytes():
    writer = OpenpyxlDetailWriter(template_path=TEMPLATE_PATH)
    filename, file_bytes = await writer.append_rows(
        [make_line_item()], year=2026, month=3, existing_data=b""
    )
    assert filename == "Chi_tiet_hoa_don_T3_2026.xlsx"
    assert isinstance(file_bytes, bytes) and len(file_bytes) > 0

@pytest.mark.asyncio
async def test_appended_row_contains_product_name():
    writer = OpenpyxlDetailWriter(template_path=TEMPLATE_PATH)
    _, file_bytes = await writer.append_rows(
        [make_line_item(ten_hang_hoa="TEST-PRODUCT")], year=2026, month=3, existing_data=b""
    )
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb.active
    all_values = [ws.cell(row=r, column=c).value for r in range(1, ws.max_row + 1) for c in range(1, 15)]
    assert "TEST-PRODUCT" in all_values

@pytest.mark.asyncio
async def test_two_appends_accumulate_rows():
    writer = OpenpyxlDetailWriter(template_path=TEMPLATE_PATH)
    _, first_bytes = await writer.append_rows(
        [make_line_item(ten_hang_hoa="Hàng A")], year=2026, month=3, existing_data=b""
    )
    _, second_bytes = await writer.append_rows(
        [make_line_item(ten_hang_hoa="Hàng B")], year=2026, month=3, existing_data=first_bytes
    )
    wb = openpyxl.load_workbook(BytesIO(second_bytes))
    ws = wb.active
    all_values = [ws.cell(row=r, column=c).value for r in range(1, ws.max_row + 1) for c in range(1, 15)]
    assert "Hàng A" in all_values
    assert "Hàng B" in all_values
